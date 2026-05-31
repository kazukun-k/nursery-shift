import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import math
import calendar
import re
from datetime import datetime
import google.generativeai as genai

# ==========================================
# 🎨 1. アプリケーション基本設定・デザイン
# ==========================================
st.set_page_config(
    page_title="保育士シフト管理システム - SmartShift",
    page_icon="📛",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 保育園にマッチした温かみと高級感のあるカスタムCSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800&family=Zen+Maru+Gothic:wght@400;500;700&display=swap');
    
    /* フォント適用 */
    html, body, [class*="css"] {
        font-family: 'Zen Maru Gothic', 'Nunito', sans-serif;
    }
    
    /* メインヘッダーのデザイン */
    .main-title {
        color: #FF6B8B;
        font-size: 2.8rem;
        font-weight: 800;
        text-align: center;
        margin-bottom: 5px;
        text-shadow: 2px 2px 4px rgba(255, 107, 139, 0.15);
    }
    .sub-title {
        color: #70A1FF;
        font-size: 1.2rem;
        text-align: center;
        margin-bottom: 30px;
        font-weight: 500;
    }
    
    /* カード風のコンポーネント */
    .card {
        background-color: #FFFFFF;
        border-radius: 15px;
        padding: 20px;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        border: 1px solid #FFE3E8;
        margin-bottom: 20px;
    }
    
    /* ボタンのカスタマイズ */
    div.stButton > button {
        background: linear-gradient(135deg, #FF6B8B 0%, #FF8E53 100%);
        color: white !important;
        border-radius: 25px;
        padding: 10px 25px;
        font-weight: 700;
        border: none;
        box-shadow: 0 4px 10px rgba(255, 107, 139, 0.3);
        transition: all 0.3s ease;
    }
    div.stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 15px rgba(255, 107, 139, 0.4);
        background: linear-gradient(135deg, #FF8E53 0%, #FF6B8B 100%);
    }
    
    /* サブヘッダーの装飾 */
    h2, h3 {
        color: #2F3542;
        border-left: 5px solid #FF6B8B;
        padding-left: 10px;
        margin-top: 20px;
    }
</style>
""", unsafe_allow_html=True)

# アプリヘッダー
st.markdown('<div class="main-title">📛 保育士シフト管理・作成システム</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">🤖 AI独自ルール解析 × 自動シフト生成エンジン搭載</div>', unsafe_allow_html=True)

# ==========================================
# 🔑 2. サイドバー設定 (APIキー & シフト基本設定)
# ==========================================
with st.sidebar:
    st.markdown("### ⚙️ システム設定")
    
    # APIキー管理
    st.subheader("🔑 Google Gemini APIキー")
    api_key_input = st.text_input(
        "Gemini APIキーを入力",
        type="password",
        placeholder="AIzaSy...",
        help="Google AI Studioで取得した無料のAPIキーを入力してください。未入力の場合は、AI機能のみ標準動作となります。"
    )
    
    if api_key_input:
        st.success("APIキーが入力されました！")
    else:
        st.warning("APIキーを入力すると、自由記述の園独自ルールや職員個人の希望休みをAIが自動解釈できるようになります。")
        
    st.markdown("""
    [👉 無料APIキーの取得手順はこちら](https://aistudio.google.com/)
    """)
    
    st.divider()
    
    # 対象年月設定
    st.subheader("📅 シフト対象年月")
    today = datetime.today()
    current_year = today.year
    current_month = today.month
    
    # 翌月をデフォルト値にする
    default_month = current_month + 1 if current_month < 12 else 1
    default_year = current_year if current_month < 12 else current_year + 1
    
    target_year = st.selectbox("対象年", range(current_year - 1, current_year + 3), index=1)
    target_month = st.selectbox("対象月", range(1, 13), index=default_month - 1)
    
    # 月の日数を計算
    num_days = calendar.monthrange(target_year, target_month)[1]
    
    st.info(f"設定期間: {target_year}年{target_month}月 ({num_days}日間)")

# ==========================================
# 🧩 3. 30分スロット定義 & 表記揺れパーサーの実装
# ==========================================
# 7:00〜20:00の30分刻み（合計 26 スロット）
TIME_SLOT_HOURS = []
for h in range(7, 20):
    TIME_SLOT_HOURS.append((h, 0, h, 30))
    TIME_SLOT_HOURS.append((h, 30, h + 1, 0))

time_slots = [f"{sh}:{sm:02d}～{eh}:{em:02d}" for sh, sm, eh, em in TIME_SLOT_HOURS]

# 正社員の固定シフトパターン (A〜E)
REGULAR_PATTERNS = {
    "A": list(range(0, 18)),  # 7:00 - 16:00
    "B": list(range(2, 20)),  # 8:00 - 17:00
    "C": list(range(4, 22)),  # 9:00 - 18:00
    "D": list(range(6, 24)),  # 10:00 - 19:00
    "E": list(range(8, 26))   # 11:00 - 20:00
}

def parse_work_hours(tz_str):
    """
    "8:30-13:00 9:00-13:30 7:00-11:00" などの複数候補をパースし、
    それぞれの候補の時間スロットインデックス(0〜25)のリストのリストを返す。
    """
    if not tz_str or not isinstance(tz_str, str):
        return []
    if tz_str == "全日":
        return [list(range(26))]
        
    s = tz_str.translate(str.maketrans({
        '０':'0', '１':'1', '２':'2', '３':'3', '４':'4', '５':'5', '６':'6', '７':'7', '８':'8', '９':'9',
        '；':':', ';':':', '：':':',
        '〜':'-', '～':'-', 'ー':'-', '－':'-', '—':'-', 'ー':'-', '─':'-'
    })).strip()
    
    parts = re.split(r'[\s,，、/]+', s)
    candidates = []
    
    for part in parts:
        if not part:
            continue
        
        match = re.search(r'(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})', part)
        if match:
            sh, sm = int(match.group(1)), int(match.group(2))
            eh, em = int(match.group(3)), int(match.group(4))
            start_idx = (sh - 7) * 2 + (1 if sm >= 30 else 0)
            end_idx = (eh - 7) * 2 + (1 if em >= 30 else 0)
            start_idx = max(0, min(25, start_idx))
            end_idx = max(0, min(26, end_idx))
            candidates.append(list(range(start_idx, end_idx)))
            continue
            
        match_h = re.search(r'(\d{1,2})\s*-\s*(\d{1,2})', part)
        if match_h:
            sh, eh = int(match_h.group(1)), int(match_h.group(2))
            start_idx = (sh - 7) * 2
            end_idx = (eh - 7) * 2
            start_idx = max(0, min(25, start_idx))
            end_idx = max(0, min(26, end_idx))
            candidates.append(list(range(start_idx, end_idx)))
            
    if not candidates:
        return [list(range(4, 20))]
    return candidates

def get_covered_slots(shift_name):
    if shift_name in REGULAR_PATTERNS:
        return REGULAR_PATTERNS[shift_name]
    elif shift_name in ["公休", "有休", "希休"]:
        return []
    candidates = parse_work_hours(shift_name)
    return candidates[0] if candidates else []

# ==========================================
# 📂 4. セッション状態の初期化
# ==========================================
if 'staff_list' not in st.session_state:
    st.session_state.staff_list = [
        {"名前": "山田 花子", "雇用形態": "正社員", "月上限日数": 20, "希望時間帯": "全日", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金,土", "off_days_json": "[]"},
        {"名前": "鈴木 一郎", "雇用形態": "パート", "月上限日数": 12, "希望時間帯": "8:30-13:00 9:00-13:30 7:00-11:00", "時短希望": "あり", "勤務可能曜日": "月,火,水,木,金", "off_days_json": '[{"day": 10, "type": "有休"}]'},
        {"名前": "佐藤 美咲", "雇用形態": "パート", "月上限日数": 15, "希望時間帯": "15:00-19:00", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金", "off_days_json": '[{"day": 15, "type": "希休"}, {"day": 16, "type": "希休"}]'},
        {"名前": "田中 恵子", "雇用形態": "正社員", "月上限日数": 20, "希望時間帯": "全日", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金,土", "off_days_json": "[]"},
        {"名前": "小林 翔太", "雇用形態": "正社員", "月上限日数": 22, "希望時間帯": "全日", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金,土", "off_days_json": "[]"},
        {"名前": "高橋 陽子", "雇用形態": "パート", "月上限日数": 16, "希望時間帯": "9:00-17:00", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金", "off_days_json": "[]"},
        {"名前": "渡辺 理恵", "雇用形態": "パート", "月上限日数": 10, "希望時間帯": "7:00-12:00 8:00-13:00", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金", "off_days_json": "[]"},
        {"名前": "伊藤 直美", "雇用形態": "パート", "月上限日数": 14, "希望時間帯": "13:30-19:00 15:00-20:00", "時短希望": "なし", "勤務可能曜日": "月,火,水,木,金", "off_days_json": "[]"},
    ]

if 'ai_rules' not in st.session_state:
    st.session_state.ai_rules = []

# 時間帯別園児数テーブルの初期化 (デフォルト在籍数でプリフィル)
if 'df_children_state' not in st.session_state:
    init_kids_data = {
        "時間帯": time_slots,
        "0歳児数": [6] * 26,
        "1-2歳児数": [12] * 26,
        "3歳児数": [18] * 26,
        "4歳以上児数": [22] * 26
    }
    st.session_state.df_children_state = pd.DataFrame(init_kids_data)

# --- 基本在籍数の変更を時間帯別テーブルに同期する関数 ---
def sync_base_kids_to_table():
    st.session_state.df_children_state["0歳児数"] = st.session_state.base_kids_0
    st.session_state.df_children_state["1-2歳児数"] = st.session_state.base_kids_1_2
    st.session_state.df_children_state["3歳児数"] = st.session_state.base_kids_3
    st.session_state.df_children_state["4歳以上児数"] = st.session_state.base_kids_4

# --- Google Gemini API 呼び出し ---
def parse_rules_with_gemini(api_key, rule_text):
    if not api_key:
        return None
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')
        
        prompt = f"""
あなたは保育園のシフト管理システムのアシスタントです。
園長が入力した以下の「園独自の特殊ルール（日本語）」を解析し、時間帯別の必要人数に対する補正ルールを抽出してください。
なお、時間帯は30分刻み（7:00から20:00まで）で計算されています。

【ルール記述】
{rule_text}

【出力フォーマット】
必ず以下の構造のJSONオブジェクトのみを返してください。不要な説明やマークダウンタグは含めず、純粋なJSONオブジェクト単体、もしくは ```json ... ``` で囲んで出力してください。
JSONには、抽出されたルールが含まれる "rules" キーのリストを設定してください。

出力スキーマ:
{{
  "rules": [
    {{
      "start_hour": 7,  // 開始時間（数値、例: 7）
      "end_hour": 9,    // 終了時間（数値、例: 9。これは9:00までを意味します）
      "min_staff": 2,   // 最低必要な保育士の人数（数値）
      "reason": "合同保育のため" // ルールの理由（文字列、例: 合同保育のため）
    }}
  ]
}}

※該当するルールがない、または解析できない場合は、空のリスト `{"rules": []}` を返してください。
"""
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        clean_text = response.text.strip()
        if clean_text.startswith("```json"):
            clean_text = clean_text[7:]
        if clean_text.endswith("```"):
            clean_text = clean_text[:-3]
        
        data = json.loads(clean_text)
        return data.get("rules", [])
    except Exception as e:
        st.error(f"Gemini APIでの解析中にエラーが発生しました: {str(e)}")
        return None

def parse_staff_details_with_gemini(api_key, remark_text):
    if not remark_text:
        return {"work_hours": "全日", "off_days": []}
        
    if not api_key:
        off_days = []
        if "10" in remark_text and "有休" in remark_text:
            off_days.append({"day": 10, "type": "有休"})
        if "15" in remark_text and "17" in remark_text and ("希望公休" in remark_text or "休み希望" in remark_text):
            off_days.extend([{"day": 15, "type": "希休"}, {"day": 16, "type": "希休"}, {"day": 17, "type": "希休"}])
            
        times = re.findall(r'\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}', remark_text)
        time_str = " ".join(times) if times else ("全日" if "全日" in remark_text else remark_text)
        
        return {"work_hours": time_str, "off_days": off_days}

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')
        
        prompt = f"""
あなたは保育園のシフト管理システムのアシスタントです。
保育士が入力した以下の「希望時間帯・備考（日本語）」を解析し、
①希望する勤務時間帯（複数ある場合はスペース区切り）
②希望する休みの日程（日付と休みの種類：有休 または 希休（希望公休））
を抽出してください。

【備考記述】
{remark_text}

【出力フォーマット】
必ず以下の構造のJSONオブジェクトのみを返してください。不要な説明やマークダウンタグは含めず、純粋なJSONオブジェクト単体、もしくは ```json ... ``` で囲んで出力してください。

出力スキーマ:
{{
  "work_hours": "8:30-13:00 9:00-13:30",  // 勤務時間の候補。時間表記以外（有休等の記述）は除外してください。全日希望の場合は "全日"
  "off_days": [
    {{
      "day": 10,       // 休みたい日付（数値のみ。例: 5月10日なら 10）
      "type": "有休"    // 休みの種類（有給休暇なら "有休"、希望公休なら "希休" のいずれか）
    }}
  ]
}}

※該当する休み希望がない場合は、"off_days": [] としてください。
"""
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        clean_text = response.text.strip()
        if clean_text.startswith("```json"):
            clean_text = clean_text[7:]
        if clean_text.endswith("```"):
            clean_text = clean_text[:-3]
        
        return json.loads(clean_text)
    except Exception as e:
        st.error(f"Gemini APIによる備考解析中にエラーが発生しました: {str(e)}")
        return {"work_hours": remark_text, "off_days": []}

# ==========================================
# タブ1: 園児数・必要人数計算
# ==========================================
with tab1:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.header("🏫 園の基本情報（在籍園児数）")
    st.caption("※ここに人数を入力すると、下の時間帯別テーブルに自動で反映されます。")
    
    col_b1, col_b2, col_b3, col_b4 = st.columns(4)
    with col_b1:
        st.number_input("0歳児クラス在籍数", min_value=0, max_value=100, value=6, key="base_kids_0", on_change=sync_base_kids_to_table)
    with col_b2:
        st.number_input("1-2歳児クラス在籍数", min_value=0, max_value=100, value=12, key="base_kids_1_2", on_change=sync_base_kids_to_table)
    with col_b3:
        st.number_input("3歳児クラス在籍数", min_value=0, max_value=100, value=18, key="base_kids_3", on_change=sync_base_kids_to_table)
    with col_b4:
        st.number_input("4歳以上児クラス在籍数", min_value=0, max_value=100, value=22, key="base_kids_4", on_change=sync_base_kids_to_table)
        
    st.divider()
    
    st.header("🏢 時間帯別の園児数調整 (30分刻み)")
    st.caption("※基本情報が自動反映されています。園児が少ない時間帯（早朝や夕方など）は、下の表をダブルクリックして手動で人数を減らしてください。")

    # 編集可能なデータエディタ
    edited_children_df = st.data_editor(
        st.session_state.df_children_state, 
        num_rows="fixed", 
        key="children_editor_v4"
    )
    st.session_state.df_children_state = edited_children_df
    
    st.subheader("📝 園独自の特殊ルール（AI自動判定）")
    custom_rule_text = st.text_area(
        "【例】「7-9時は全学年合同で過ごすため最低2人配置」「17:00以降は遅番で最低2人必要」など、独自の配置制限を文章で記述してください。",
        value="7-9時は全学年合同で過ごすため最低2人配置してください。また、17-19時は合同で最低2人配置にしてください。",
        placeholder="ここに園独自のルールを自由に文章で入力してください。"
    )
    
    col_ai_btn, col_ai_status = st.columns([1, 3])
    with col_ai_btn:
        run_ai = st.button("AIルールを解析して反映する")
        
    if run_ai:
        if not api_key_input:
            st.warning("⚠️ APIキーが設定されていません。無料デモ用ルールを適用します。")
            if "7-9" in custom_rule_text and "合同" in custom_rule_text:
                st.session_state.ai_rules = [
                    {"start_hour": 7, "end_hour": 9, "min_staff": 2, "reason": "【デモ適用】7-9時合同保育のため最低2名"}
                ]
                if "17-19" in custom_rule_text or "17時" in custom_rule_text:
                    st.session_state.ai_rules.append(
                        {"start_hour": 17, "end_hour": 19, "min_staff": 2, "reason": "【デモ適用】17-19時合同保育のため最低2名"}
                    )
            st.success("デモルールを適用しました。")
        else:
            with st.spinner("Gemini AIがルールを解析中..."):
                extracted_rules = parse_rules_with_gemini(api_key_input, custom_rule_text)
                if extracted_rules:
                    st.session_state.ai_rules = extracted_rules
                    st.success(f"🎉 AI解析成功！ {len(extracted_rules)}件 of ルールを適用しました。")
                else:
                    st.info("特殊ルールは検出されませんでした。標準の配置基準を適用します。")
                    st.session_state.ai_rules = []
                    
    if st.session_state.ai_rules:
        st.info("💡 **現在適用中のAI補正ルール:**\n" + \
                "\n".join([f"- **{r['start_hour']}:00～{r['end_hour']}:00**: 最低 {r['min_staff']}人配置 ({r['reason']})" for r in st.session_state.ai_rules]))

    st.divider()

    # 配置基準＋AI補正計算
    def calculate_required_staff_with_ai(row):
        current_hour = int(row["時間帯"].split(":")[0])
        
        count = (row["0歳児数"]/3) + (row["1-2歳児数"]/6) + (row["3歳児数"]/20) + (row["4歳以上児数"]/30)
        base_staff = math.ceil(count)
        
        min_staff_req = 1
        for rule in st.session_state.ai_rules:
            if rule["start_hour"] <= current_hour < rule["end_hour"]:
                min_staff_req = max(min_staff_req, rule["min_staff"])
                
        final_staff = max(base_staff, min_staff_req)
        return final_staff

    edited_children_df["必要保育士数"] = edited_children_df.apply(calculate_required_staff_with_ai, axis=1)
    
    st.subheader("⏰ 計算結果：時間帯別の必要保育士数")
    result_df = edited_children_df[["時間帯", "必要保育士数"]].set_index("時間帯").T
    st.dataframe(result_df)
    st.markdown('</div>', unsafe_allow_html=True)

# ==========================================
# タブ2: 職員条件設定
# ==========================================
with tab2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.header("👥 保育士・スタッフの勤務条件設定")
    
    type_staff = st.selectbox("雇用形態", ["正社員", "パート"], key="reg_type_staff")
    
    with st.form("add_staff_form"):
        st.subheader("➕ 職員の新規登録")
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("お名前", placeholder="例：山田 花子")
            max_days = st.number_input("月間勤務上限日数", min_value=1, max_value=31, value=20 if type_staff == "正社員" else 12)
            
            default_days = ["月", "火", "水", "木", "金", "土"] if type_staff == "正社員" else ["月", "火", "水", "木", "金"]
            available_days = st.multiselect(
                "勤務可能曜日（チェック項目）", 
                ["月", "火", "水", "木", "金", "土", "日"], 
                default=default_days
            )
        with col2:
            remark = st.text_area(
                "希望時間帯・休み等の備考 (自由記述)", 
                value="全日" if type_staff == "正社員" else "8:30-13:00 9:00-13:30。10日は有休、15-17日は希望公休", 
                placeholder="例: 8:30-13:00。10日は有休、15日は希望公休",
                help="【AI解析対象】勤務可能時間の候補と、休みたい日付（例: 10日は有休、15日は希望公休）を文章で自由に書きます。"
            )
            short_time = st.selectbox("時短希望", ["なし", "あり"])
            
            submit = st.form_submit_button("🔍 AIで備考を解析して職員を追加")
            
        if submit and name:
            with st.spinner("AIが備考欄から希望時間と希望休みを解析中..."):
                parsed = parse_staff_details_with_gemini(api_key_input, remark)
                days_str = ",".join(available_days)
                
                st.session_state.staff_list.append({
                    "名前": name, 
                    "雇用形態": type_staff, 
                    "月上限日数": max_days, 
                    "希望時間帯": parsed["work_hours"], 
                    "時短希望": short_time,
                    "勤務可能曜日": days_str,
                    "off_days_json": json.dumps(parsed["off_days"])
                })
                st.success(f"✅ {name} さんを登録しました！\n"
                           f"- 解析された希望時間: {parsed['work_hours']}\n"
                           f"- 解析された希望休み: {parsed['off_days']}")

    st.divider()

    st.subheader("📋 現在の職員一覧")
    st.caption("※一覧のデータをダブルクリックで修正できます。休み希望データは「off_days_json」列で管理されています。")
    
    df_staff = pd.DataFrame(st.session_state.staff_list)
    edited_staff_df = st.data_editor(df_staff, num_rows="dynamic", key="staff_editor")
    
    if st.button("変更を確定して職員リストを保存"):
        st.session_state.staff_list = edited_staff_df.to_dict('records')
        st.success("✅ 職員リストの変更を確定しました。")
    
    if st.button("⚠️ 一覧をクリアして初期化"):
        st.session_state.staff_list = []
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# ==========================================
# タブ3: シフト自動生成・Excel出力
# ==========================================
with tab3:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.header("🗓️ シフト自動生成エンジン")
    st.write(f"**対象期間: {target_year}年{target_month}月** ({num_days}日間)")
    
    # 凡例の表示
    st.info("""
    📖 **シフト記号・休みの凡例:**
    - **A**: 7:00 - 16:00 (正社員)
    - **B**: 8:00 - 17:00 (正社員)
    - **C**: 9:00 - 18:00 (正社員)
    - **D**: 10:00 - 19:00 (正社員)
    - **E**: 11:00 - 20:00 (正社員)
    - **有休**: 有給休暇 (勤務日数から除外、AI解析で自動設定)
    - **希休**: 希望公休 (勤務日数から除外、AI解析で自動設定)
    - **公休**: 通常の公休
    *(※パート職員は、希望の複数時間から最適な時間が自動的に選択されます。)*
    """)
    
    # --- シフト生成アルゴリズム ---
    def generate_auto_schedule(df_kids, staff_list, target_year, target_month, num_days):
        req_by_slot = {}
        for idx, row in df_kids.iterrows():
            req_by_slot[idx] = row["必要保育士数"]
            
        staff_work_days = {staff["名前"]: 0 for staff in staff_list}
        
        days_columns = [f"{d}日" for d in range(1, num_days + 1)]
        schedule_data = []
        for staff in staff_list:
            row = {"名前": staff["名前"], "雇用形態": staff["雇用形態"]}
            for col in days_columns:
                row[col] = "公休"
            schedule_data.append(row)
            
        schedule_df = pd.DataFrame(schedule_data)
        schedule_df.set_index("名前", inplace=True)
        
        work_history = {staff["名前"]: [False] * (num_days + 1) for staff in staff_list}
        
        # 1. 休み希望（有休・希休）を最初に入力・確定させる
        for s in staff_list:
            s_name = s["名前"]
            off_days_data = json.loads(s.get("off_days_json", "[]"))
            for off in off_days_data:
                day_num = off["day"]
                off_type = off["type"]
                if 1 <= day_num <= num_days:
                    schedule_df.at[s_name, f"{day_num}日"] = off_type

        # 2. 日ごとの勤務割り当て処理
        for d in range(1, num_days + 1):
            day_col = f"{d}日"
            weekday = calendar.weekday(target_year, target_month, d)
            is_weekend = weekday in [5, 6]
            
            weekday_map = ["月", "火", "水", "木", "金", "土", "日"]
            day_of_week = weekday_map[weekday]
            
            daily_req = {}
            for idx in range(26):
                req = req_by_slot.get(idx, 1)
                if is_weekend:
                    req = max(1, math.ceil(req * 0.5))
                daily_req[idx] = req
                
            assigned_today = {}
            assigned_counts = {idx: 0 for idx in range(26)}
            
            for s in staff_list:
                s_name = s["名前"]
                existing_shift = schedule_df.at[s_name, day_col]
                if existing_shift in ["有休", "希休"]:
                    assigned_today[s_name] = existing_shift
            
            def get_eligible_candidates(assigned_today):
                candidates = []
                for s in staff_list:
                    s_name = s["名前"]
                    if s_name in assigned_today:
                        continue
                    if staff_work_days[s_name] >= s["月上限日数"]:
                        continue
                        
                    # 勤務可能曜日のチェック
                    available_days_str = s.get("勤務可能曜日", "月,火,水,木,金,土")
                    available_days_list = [x.strip() for x in available_days_str.split(",") if x.strip()]
                    if day_of_week not in available_days_list:
                        continue
                        
                    # 勤務ペース管理（連勤・ペース配分）
                    start_day = max(1, d - 6)
                    recent_work_days = sum(work_history[s_name][start_day:d])
                    weekly_pace_limit = math.ceil(s["月上限日数"] * 7 / 30)
                    
                    consecutive_days = 0
                    check_day = d - 1
                    while check_day >= 1 and work_history[s_name][check_day]:
                        consecutive_days += 1
                        check_day -= 1
                    
                    over_pace = recent_work_days >= weekly_pace_limit
                    too_many_consecutive = consecutive_days >= 5
                    
                    candidates.append({
                        "staff": s,
                        "work_days": staff_work_days[s_name],
                        "recent_work": recent_work_days,
                        "consecutive": consecutive_days,
                        "over_pace": over_pace,
                        "too_many_consecutive": too_many_consecutive
                    })
                return candidates

            # 割り当てループ
            while True:
                understaffed_slots = []
                for idx in range(26):
                    deficit = daily_req[idx] - assigned_counts[idx]
                    if deficit > 0:
                        understaffed_slots.append((idx, deficit))
                        
                if not understaffed_slots:
                    break
                    
                candidates = get_eligible_candidates(assigned_today)
                if not candidates:
                    break
                    
                understaffed_slots.sort(key=lambda x: x[1], reverse=True)
                target_slot = understaffed_slots[0][0]
                
                best_match = None
                best_score = -999999
                best_pattern = None
                
                for cand in candidates:
                    s = cand["staff"]
                    role = s["雇用形態"]
                    patterns = []
                    
                    if role == "正社員":
                        if s["希望時間帯"] == "全日":
                            for pat, slots in REGULAR_PATTERNS.items():
                                patterns.append((pat, slots))
                        else:
                            custom_slots_list = parse_work_hours(s["希望時間帯"])
                            for slots in custom_slots_list:
                                patterns.append((s["希望時間帯"], slots))
                    else: # パート
                        custom_slots_list = parse_work_hours(s["希望時間帯"])
                        parts = re.split(r'[\s,，、/]+', s["希望時間帯"])
                        for i, slots in enumerate(custom_slots_list):
                            label = parts[i] if i < len(parts) else s["希望時間帯"]
                            patterns.append((label, slots))
                            
                    for pat_name, slots in patterns:
                        if target_slot not in slots:
                            continue
                            
                        overlap = sum(1 for sl in slots if daily_req[sl] > assigned_counts[sl])
                        
                        score = overlap * 10
                        score += (s["月上限日数"] - cand["work_days"]) * 5
                        score -= cand["recent_work"] * 3
                        score -= cand["consecutive"] * 2
                        
                        if cand["over_pace"]:
                            score -= 15
                        if cand["too_many_consecutive"]:
                            score -= 30
                            
                        if score > best_score:
                            best_score = score
                            best_match = cand
                            best_pattern = (pat_name, slots)
                            
                if best_match and best_pattern:
                    s_name = best_match["staff"]["名前"]
                    pat_name, slots = best_pattern
                    
                    assigned_today[s_name] = pat_name
                    staff_work_days[s_name] += 1
                    work_history[s_name][d] = True
                    for sl in slots:
                        assigned_counts[sl] += 1
                else:
                    if len(understaffed_slots) > 1:
                        target_slot = understaffed_slots[1][0]
                        break
                    else:
                        break
                        
            for s_name, shift in assigned_today.items():
                if schedule_df.at[s_name, day_col] not in ["確実に休み", "有休", "希休"]:
                    schedule_df.at[s_name, day_col] = shift
                
        schedule_df.reset_index(inplace=True)
        return schedule_df

    # シフト生成ボタン
    if len(st.session_state.staff_list) == 0:
        st.warning("⚠️ 職員が登録されていません。")
    else:
        if st.button("⚡ シフトを自動生成する"):
            with st.spinner("AI休み希望と配置要件に基づき、最適なシフトを計算中..."):
                schedule_result = generate_auto_schedule(
                    st.session_state.df_children_state, 
                    st.session_state.staff_list, 
                    target_year, 
                    target_month, 
                    num_days
                )
                st.session_state.schedule_result = schedule_result
                st.success("🎉 シフト表を自動作成しました！")

        # 生成結果があれば表示
        if 'schedule_result' in st.session_state:
            st.subheader("📅 作成されたシフト表 (手動微調整も可能)")
            st.caption("※手動でセルを修正すると、下の人員不足チェッカーもリアルタイムで連動します。")
            
            edited_schedule = st.data_editor(
                st.session_state.schedule_result, 
                key="schedule_editor", 
                num_rows="fixed"
            )
            
            if st.button("シフト表の調整内容を保存"):
                st.session_state.schedule_result = edited_schedule
                st.success("✅ 調整されたシフトを保存しました！")

            # ==========================================
            # 🔍 リアルタイム人員不足チェッカー
            # ==========================================
            st.divider()
            st.subheader("🔍 時間帯別の人員不足チェッカー")
            
            shortages = []
            req_by_slot = {}
            for idx, row in st.session_state.df_children_state.iterrows():
                req_by_slot[idx] = row["必要保育士数"]
                
            for d in range(1, num_days + 1):
                day_col = f"{d}日"
                weekday = calendar.weekday(target_year, target_month, d)
                is_weekend = weekday in [5, 6]
                
                weekday_map = ["月", "火", "水", "木", "金", "土", "日"]
                day_of_week = weekday_map[weekday]
                
                daily_req = {}
                for idx in range(26):
                    req = req_by_slot.get(idx, 1)
                    if is_weekend:
                        req = max(1, math.ceil(req * 0.5))
                    daily_req[idx] = req
                    
                assigned_counts = {idx: 0 for idx in range(26)}
                for idx_row, row in edited_schedule.iterrows():
                    shift = row[day_col]
                    if shift not in ["公休", "有休", "希休"]:
                        for sl in get_covered_slots(shift):
                            assigned_counts[sl] += 1
                            
                # 不足スロットを検出
                for idx in range(26):
                    deficit = daily_req[idx] - assigned_counts[idx]
                    if deficit > 0:
                        shortages.append({
                            "day": d,
                            "day_col": day_col,
                            "weekday": day_of_week,
                            "slot_name": time_slots[idx],
                            "required": daily_req[idx],
                            "assigned": assigned_counts[idx],
                            "deficit": deficit
                        })

            if shortages:
                st.error("⚠️ **保育士の配置不足が発生している時間帯があります（下の情報を元に公休をずらす等の交渉をしてください）**")
                
                shortages_df = pd.DataFrame(shortages)
                for day, group in shortages_df.groupby("day_col"):
                    weekday_str = group.iloc[0]["weekday"]
                    msg = f"**📅 {day}({weekday_str}) の人員不足箇所:**\n"
                    for _, r in group.iterrows():
                        msg += f"- `{r['slot_name']}` : **{r['deficit']}名 不足** (必要: {r['required']}名 / 配置: {r['assigned']}名)\n"
                    st.warning(msg)
            else:
                st.success("🎉 おめでとうございます！すべての日程・時間帯で必要人数が完全に確保されています！")

            # 統計情報の表示
            st.subheader("📊 職員ごとの出勤日数カウント")
            stat_data = []
            for idx, row in edited_schedule.iterrows():
                work_days_count = sum(1 for col in edited_schedule.columns if "日" in col and row[col] not in ["公休", "有休", "希休"])
                stat_data.append({
                    "名前": row["名前"],
                    "雇用形態": row["雇用形態"],
                    "当月出勤日数": work_days_count
                })
            st.dataframe(pd.DataFrame(stat_data))

            # ==========================================
            # Excel出力セクション
            # ==========================================
            st.divider()
            st.subheader("💾 完成したシフト表をダウンロード")
            
            def generate_styled_excel(df_kids, df_teachers, df_schedule, year, month):
                wb = openpyxl.Workbook()
                
                font_title = Font(name="Meiryo UI", size=14, bold=True, color="2F3542")
                font_header = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
                font_data = Font(name="Meiryo UI", size=9)
                font_bold = Font(name="Meiryo UI", size=9, bold=True)
                font_legend = Font(name="Meiryo UI", size=9, italic=True, color="57606F")
                
                fill_primary = PatternFill(start_color="FF6B8B", end_color="FF6B8B", fill_type="solid")
                fill_secondary = PatternFill(start_color="70A1FF", end_color="70A1FF", fill_type="solid")
                fill_weekend = PatternFill(start_color="F1F2F6", end_color="F1F2F6", fill_type="solid")
                fill_off = PatternFill(start_color="E4E7EB", end_color="E4E7EB", fill_type="solid")
                fill_paid = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                fill_requested = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                border_thin = Border(
                    left=Side(style='thin', color='CED6E0'),
                    right=Side(style='thin', color='CED6E0'),
                    top=Side(style='thin', color='CED6E0'),
                    bottom=Side(style='thin', color='CED6E0')
                )
                
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")
                
                ws_schedule = wb.active
                ws_schedule.title = "シフト表"
                ws_schedule.views.sheetView[0].showGridLines = True
                
                ws_schedule.cell(row=1, column=1, value=f"📛 {year}年{month}月 保育士シフト表").font = font_title
                ws_schedule.row_dimensions[1].height = 30
                
                headers = list(df_schedule.columns)
                ws_schedule.append([])
                ws_schedule.append(headers + ["出勤日数"])
                ws_schedule.row_dimensions[3].height = 25
                
                for c_idx in range(1, len(headers) + 2):
                    cell = ws_schedule.cell(row=3, column=c_idx)
                    cell.fill = fill_primary
                    cell.font = font_header
                    cell.alignment = align_center
                    cell.border = border_thin
                
                last_row_idx = 3
                for r_idx, row in enumerate(df_schedule.values, 4):
                    row_list = list(row)
                    ws_schedule.append(row_list)
                    ws_schedule.row_dimensions[r_idx].height = 22
                    last_row_idx = r_idx
                    
                    last_col_letter = get_column_letter(len(headers))
                    formula = f'=COUNTIFS(C{r_idx}:{last_col_letter}{r_idx}, "<>公休", C{r_idx}:{last_col_letter}{r_idx}, "<>有休", C{r_idx}:{last_col_letter}{r_idx}, "<>希休")'
                    ws_schedule.cell(row=r_idx, column=len(headers) + 1, value=formula).font = font_bold
                    ws_schedule.cell(row=r_idx, column=len(headers) + 1).alignment = align_center
                    ws_schedule.cell(row=r_idx, column=len(headers) + 1).border = border_thin
                    
                    for c_idx in range(1, len(headers) + 1):
                        cell = ws_schedule.cell(row=r_idx, column=c_idx)
                        cell.font = font_data
                        cell.border = border_thin
                        
                        if c_idx == 1:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center
                            
                        val = cell.value
                        if val == "公休":
                            cell.fill = fill_off
                            cell.font = Font(name="Meiryo UI", size=9, color="747D8C")
                        elif val == "暗黒":
                            cell.fill = fill_off
                            cell.font = Font(name="Meiryo UI", size=9, color="747D8C")
                        elif val == "有休":
                            cell.fill = fill_paid
                            cell.font = Font(name="Meiryo UI", size=9, bold=True, color="C00000")
                        elif val == "希休":
                            cell.fill = fill_requested
                            cell.font = Font(name="Meiryo UI", size=9, bold=True, color="7F6000")
                        elif val in ["A", "B", "C", "D", "E"]:
                            cell.font = Font(name="Meiryo UI", size=9, bold=True)
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        elif val:
                            cell.font = Font(name="Meiryo UI", size=9, bold=True)
                            
                for col in ws_schedule.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws_schedule.column_dimensions[col_letter].width = max(max_len + 3, 11)
                
                # シフト記号の凡例
                legend_start_row = last_row_idx + 3
                ws_schedule.cell(row=legend_start_row, column=1, value="📖 シフト記号・休暇 凡例:").font = font_bold
                legend_items = [
                    ("A", "7:00 - 16:00 (正社員)"),
                    ("B", "8:00 - 17:00 (正社員)"),
                    ("C", "9:00 - 18:00 (正社員)"),
                    ("D", "10:00 - 19:00 (正社員)"),
                    ("E", "11:00 - 20:00 (正社員)"),
                    ("有休", "有給休暇 (出勤日数カウント対象外)"),
                    ("希休", "希望公休 (出勤日数カウント対象外)")
                ]
                for idx, (sym, tm) in enumerate(legend_items):
                    row_num = legend_start_row + 1 + idx
                    ws_schedule.cell(row=row_num, column=1, value=f"  {sym} : {tm}").font = font_legend
                
                # --- シート2: 必要人数要件 ---
                ws_kids = wb.create_sheet(title="時間帯別必要人数")
                ws_kids.views.sheetView[0].showGridLines = True
                
                ws_kids.cell(row=1, column=1, value="📊 時間帯別 園児数と必要人数").font = font_title
                ws_kids.append([])
                
                for r_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(df_kids, index=False, header=True), 3):
                    ws_kids.append(row)
                    ws_kids.row_dimensions[r_idx].height = 20
                    for c_idx in range(1, len(row) + 1):
                        cell = ws_kids.cell(row=r_idx, column=c_idx)
                        cell.border = border_thin
                        cell.alignment = align_center
                        if r_idx == 3:
                            cell.fill = fill_secondary
                            cell.font = font_header
                        else:
                            cell.font = font_data
                
                for col in ws_kids.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws_kids.column_dimensions[col_letter].width = max(max_len + 5, 12)
                
                # --- シート3: 職員マスター ---
                ws_staff = wb.create_sheet(title="職員マスター")
                ws_staff.views.sheetView[0].showGridLines = True
                
                ws_staff.cell(row=1, column=1, value="👥 職員（スタッフ）勤務条件一覧").font = font_title
                ws_staff.append([])
                
                for r_idx, row in enumerate(openpyxl.utils.dataframe.dataframe_to_rows(df_teachers, index=False, header=True), 3):
                    ws_staff.append(row)
                    ws_staff.row_dimensions[r_idx].height = 20
                    for c_idx in range(1, len(row) + 1):
                        cell = ws_staff.cell(row=r_idx, column=c_idx)
                        cell.border = border_thin
                        cell.alignment = align_center
                        if r_idx == 3:
                            cell.fill = fill_secondary
                            cell.font = font_header
                        else:
                            cell.font = font_data
                            if c_idx == 1:
                                cell.alignment = align_left
                
                for col in ws_staff.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws_staff.column_dimensions[col_letter].width = max(max_len + 5, 15)

                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer

            excel_data = generate_styled_excel(st.session_state.df_children_state, df_staff, edited_schedule, target_year, target_month)
            
            st.download_button(
                label="📥 完成したシフト管理Excelファイルをダウンロード",
                data=excel_data,
                file_name=f"保育園シフト管理表_{target_year}年{target_month}月.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    st.markdown('</div>', unsafe_allow_html=True)
